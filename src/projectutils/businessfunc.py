""" Busiess Functions 
    src/projectutils/businessfunc.py 
    This file contains the functions related to the main business functions.
    Author: vipulasrilanka@yahoo.com 
    (c) 2024 """

import openpyxl
import os
import re
from num2words import num2words
from datetime import datetime

from constants.errorcodes import ERROR_SUCCESS, ERROR_UNKNOWN, ERROR_NULL_STRING, ERROR_FILE_NOT_FOUND
from constants.templatedata import REC_COL_INDEX, REC_COL_KEY, REC_COL_STR_ID
from constants.templatedata import TEMP_COL_INDEX, TEMP_COL_NAME, TEMP_COL_CONTENT, TEMP_COL_PARAM, TEMP_COL_PRE_PROC
from constants.templatedata import TEMP_MIN_STR_DATA_LENGTH

def getStringFromFileObject(fileName,fileOjectList,fileSheetName,primeryKey,primeryKeyCol,valueCol):
    """Get the designated text from excel file object. It will look for the file name in the
       fileOjectList["name"] and get the file object from fileOjectList["object"]. Open the fileSheetName
       from the object, and search the primeryKeyCol for primeryKey. If it matches, will return the value 
       in the valueCol. 
    Args: fileName (string) : name of the excel workbook
          fileOjectList (list) : directory withfile object against the name
          fileSheetName (string) : sheet name in the workbook
          primeryKey (string/number) : matching condition to look for
          primeryKeyCol (string) : the column ID to match
          valueCol (string) : the column ID to return data from
    Returns string: relevent text from the file, or None """
    #check for argumanet validity and rerurn if there are errors <TODO>
    #go through each file object and look for a match.
    for sourceFile in fileOjectList:
        if(fileName == sourceFile["name"]):
            workBook = sourceFile["object"]
            print("extract record id ["+str(primeryKey) +"] from ["+fileName+ "]")
            sheet = workBook[fileSheetName]
            # Convert the primaryKeyCol and valueCol to numeric indices
            primeryKeyColIndex = openpyxl.utils.column_index_from_string(primeryKeyCol)
            valueColIndex = openpyxl.utils.column_index_from_string(valueCol)
            print("primeryKeyColIndex ["+ str(primeryKeyColIndex) + "] valueColIndex [" + str(valueColIndex) + "]")

            # Loop through the rows, starting from the second row (assuming row 1 is the header)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                # Check if the value in the primaryKeyCol matches the given primaryKey
                matchString = str(row[primeryKeyColIndex - 1].value)
                print ("Try: match", matchString, "with", primeryKey)
                if matchString == str(primeryKey):
                    # Return the value from the valueCol in the matching row
                    stringValue = str(row[valueColIndex - 1].value)
                    print("Found", matchString, " => ", stringValue)
                    return stringValue
            # If the primary key isn't found, return Error
            print("ERROR: Can not find primery key")
            return ERROR_NULL_STRING

def concatString(pdfOverlayList,overlayName,overlayString):
    """ Process string concatnation. This function will get a concantation logic,
        find the matching overly item and add the new string at the end of the
        overlay text.
    Args: pdfOverlayList (list) directory that contains the current overlay list 
          overlayName (string) name of the overlay to add "overlayString"
          overlayString (string) the text to add.
    Returns: int: Error code """
    # Define the regex pattern to match the format !<CONCAT><STRING>
    pattern = r"^!<CONCAT><(.+)>$"
    # Use re.match to check if the input string matches the pattern
    match = re.match(pattern, overlayName)
    if match:
        # Extract the text between the second <>
        exString = match.group(1)
    else:
        # If the format is incorrect
        print("ERROR: The input string does not have the correct format: !<CONCAT><STRING>")
        return ERROR_UNKNOWN
    for pdfOverlay in pdfOverlayList:
        if pdfOverlay["name"] == exString:
            #found the matching location
            print ("Concat ["+ overlayString + "] to " + pdfOverlay["name"])
            pdfOverlay["string"] = str(pdfOverlay["string"]) + str(overlayString)
    return ERROR_SUCCESS

def loadTemplateData(templateFile,sheetName):
    """ Reads the text overlays from the template file.
    Args: templateFile (string): The template file name
          sheetName (string): The sheet name with data
    Returns: list: A list of ordered dictionaries, or error code """

    print("+ Fn: loadTemplateData")
    # variable to return data
    textOverlayList = [] 
    #check if the file path is valid
    if( not os.path.exists(templateFile)):
        print("Error [loadTemplateData]: Tempate file not found")
        return ERROR_FILE_NOT_FOUND # error 
    #open template file, and load the sheet
    templateBook = openpyxl.load_workbook(templateFile)
    textOverlayDataSheet = templateBook[sheetName]
    print("Debug [loadTemplateData]: Sheet Size",textOverlayDataSheet.dimensions, textOverlayDataSheet.max_row, " Rows ", textOverlayDataSheet.max_column, " columns")
    #go through every text overlay item
    for overlays in textOverlayDataSheet.rows:
        #Get the index to a string. Empty cells will be string "None"
        rowIndex = str(overlays[TEMP_COL_INDEX].value).strip() #ignore space
        # stop if we reach an empty cell
        if("None" == rowIndex):
            break
        # the index has to be always numbers, skip others
        if( not rowIndex.isdigit()):
            print("Warning [loadTemplateData]: skip Row Index ["+rowIndex+ "]")
            continue
        content = str(overlays[TEMP_COL_CONTENT].value).strip() #ignore space
        #user initiated end of loop.
        if("None" == content):
            print("Warning [loadTemplateData]: User terminated at Index ["+rowIndex+"]")
            break
        if(not (content.startswith('<') and content.endswith('>') and len(content) > TEMP_MIN_STR_DATA_LENGTH)):
            print("Error [loadTemplateData]: Data Error at Index ["+rowIndex+"]")
            break
        param = validateParams(str(overlays[TEMP_COL_PARAM].value).strip())
        preprocess = validateParams(str(overlays[TEMP_COL_PRE_PROC].value).strip())
        content = validateParams(content)
        textOverlayList.append({"name": str(overlays[TEMP_COL_NAME].value).strip(), 
                                "content": content,
                                "param": param,
                                "preProcess": preprocess})
        #check the item 2, File locked
        if "Text" == content["Type"]:
            print(rowIndex, "overlay type > immidiate text =>", content["Text"])
            #Save immidiate text data
        elif "File" == content["Type"]:
            print(rowIndex, "overlay type > From file => ", content["File"])
        else:
            print(rowIndex, "Error: undefined overlay type : ", content["type"])
            break
    # Data store is done. return
    print("- Fn: loadTemplateData")
    return textOverlayList

def validateParams(paramString):
    """ Get the param data from the file and break it down to param list.
        return: None if there is no param data."""
    print("+Fn validateParams (", paramString, ")")
    params = {}
    # Use a regular expression to extract all the <key=value> pairs
    pattern = r"<(.*?)=(.*?)>"
    matches = re.findall(pattern, paramString)
    if len(matches) == 0:
        return None
    # Iterate over each match and add it to the dictionary
    for key, value in matches:
        # Try to convert value to int or float, if applicable
        value = extractValueFromString(value)
        value = convertFunctionString(value)
        params[key] = value  # Keep the original string if conversion fails
    return params


def getFilesFromOverlayList(textOverlayList):
    """ Returns a unique list of file names in the overlay 
        Note: having no files in the list is not an error.
    Args: textOverlayList (list): list of directories
    Returns: list: A list of file names, strings """

    print("+ Fn: getFilesFromOverlayList")
    fileNameList = []
    #go through each overlay
    for textOverlay in textOverlayList:
        if "File" == textOverlay["content"]["Type"]:
            # there is a file attribute
            filename = textOverlay["content"]["File"]
            print("Fould a file ", filename)
            #check if this file is already in the list
            if filename not in fileNameList:
                fileNameList.append(filename)
    print("- Fn: getFilesFromOverlayList")
    return fileNameList


def loadRecordIdList(templateFile,sheetName):
    """ loadRecordIdList: This will load the records to process 
    Args:   templateFile (string): Template file name
            sheetName (string): The sheet name with data
    Returns: list: a list of directories with keys and identifiers of records or error code """

    print("+ Fn: loadRecordIdList")
    # variable to return data
    recordIdList = [] 
    #check if the file path is valid
    if( not os.path.exists(templateFile)):
        print("Error [loadRecordIdList]: Tempate file not found")
        return ERROR_FILE_NOT_FOUND # error code
    #open template file, and load the sheet
    templateBook = openpyxl.load_workbook(templateFile, data_only=True)
    recordIdDataSheet = templateBook[sheetName]
    print("Debug [loadRecordIdList]: Sheet Size",recordIdDataSheet.dimensions, recordIdDataSheet.max_row, " Rows ", recordIdDataSheet.max_column, " columns")
    #go through every text overlay item
    for record in recordIdDataSheet.rows:
        #Get the index to a string. Empty cells will be string "None"
        rowIndex = str(record[REC_COL_INDEX].value)
        print(rowIndex)
        # stop if we reach an empty cell
        if("None" == rowIndex):
            break
        # the index has to be always numbers, skip others
        if( not rowIndex.isdigit()):
            print("Warning [loadRecordIdList]: skip Row Index : ", rowIndex)
            continue
        primeryKey = str(record[REC_COL_KEY].value)
        #user initiated end of loop.
        if("None" == primeryKey):
            print("Warning [loadRecordIdList]: User terminated at Index : ", rowIndex)
            break
        if not primeryKey.isdigit():
            print("Error [primeryKey]: Data Error at Index : ",rowIndex)
            break
        recordIdList.append({"key": int(primeryKey), "identifier": str(record[REC_COL_STR_ID].value)})
    return recordIdList

def getNumber(text, type):
    """ getNumber extracts a numeric value from a given text string. """
    if isinstance(text,int) or isinstance(text,float):
        print("Fn getNumber: input is a number")
    elif isinstance(text,str):
        match = re.search(r"[\d,]+\.\d+|[\d,]+", str(text))    
        if match:
            # Remove commas and convert to float
            text = float(match.group().replace(',', ''))
        else:
            text = 0
    return type(text)

def preprocess(text,processList):
    #{'function': {'name': 'AddSpace', 'param1': 'None'}}
    """ Preprocess text string based on the process given"""
    print("+Fn preprocess Text =>[", text,"]", processList)
    if "Function" in processList:
        if "NumberToText" == processList["Function"]["name"]:
            number = None
            if "Integer" == processList["Function"]["param2"]:
                #round off to integer
                number = getNumber(text,round)
            elif "Floating Point" == processList["Function"]["param2"]:
                number = getNumber(text,float)
            else:
                #default
                number = getNumber(text,float)
            #convert the number to text
            number = num2words(number,to = 'cardinal')
            #make the first letter capital, and return as text.
            return number.capitalize()
        
        elif "AddSpace" == processList["Function"]["name"]:
            count = getNumber(processList["Function"]["param2"],int)
            return str(text)+ (" " * count)
        elif "FormatDate" == processList["Function"]["name"]:
            print("[preprocess][FormatDate][param2]", processList["Function"]["param2"], type(processList["Function"]["param2"]))
            return formatDate(text,processList["Function"]["param2"])
        elif "NumberToCurrency" == processList["Function"]["name"]:
            print("[preprocess][NumberToCurrency][param2]", processList["Function"]["param2"], type(processList["Function"]["param2"]))
            print("[preprocess][NumberToCurrency][param3]", processList["Function"]["param3"], type(processList["Function"]["param3"]))
            return getCurrencyString(text,processList["Function"]["param2"],processList["Function"]["param3"])
        elif "FormatNumber" == processList["Function"]["name"]:
            print("[preprocess][FormatNumber][param2]", processList["Function"]["param2"], type(processList["Function"]["param2"]))
            print("[preprocess][FormatNumber][param3]", processList["Function"]["param3"], type(processList["Function"]["param3"]))
            print("[preprocess][FormatNumber][param4]", processList["Function"]["param4"], type(processList["Function"]["param4"]))
            return getFormattedNumber(text,processList["Function"]["param2"],processList["Function"]["param3"],processList["Function"]["param4"])
        elif "changeTextCase" == processList["Function"]["name"]:
            caseType = processList["Function"]["param2"]
            print("[preprocess][changeTextCase][param2]", caseType, type(caseType))
            if("lower" == caseType):
                return text.lower()
            elif ("UPPER" == caseType):
                return text.upper()
            elif ("Title" == caseType):
                return text.title()
            else:
                print("unsupported Case", caseType)
                return text
        else:
            print("Error [preprocess] Unsupported Pre-process function")
    return text

def formatDate(date_string,format):
    #$ print(date_string,format)
    # check the type of input
    if not isinstance(date_string,str): return ("Data Type Error")
    if not isinstance(format,str): format = "%Y/%m/%d" #set to default
    # try formats. We have added extra formats to cover data objects coming from openpyxl
    formats = [format, "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]
    parsed_date = None
    #try each format and see if the date fits any. Load it to a date time object
    for fmt in formats:
        try:
            parsed_date = datetime.strptime(date_string, fmt)
            break
        except ValueError:
            continue
    if parsed_date is None:
        # the text can not be decoded.
        print("No matching format found for:", date_string, "format =", format)
        return date_string
    # we have a valid date. Return the formatted string
    # print(parsed_date.strftime(format))
    return (parsed_date.strftime(format))    

def getFormattedNumber(number, decimalPoints, prefix, suffix):
    """Returns a text string with the prefix and suffix"""
    number = getNumber(number,float)
    # Format the number to the specified number of decimal places
    number = f"{number:,.{decimalPoints}f}"   
    # Return the currency with formatted number
    return f"{prefix}{number}{suffix}"

def getCurrencyString(number,decimalPoints,currency):
    """Returns a text string with the number in currency format"""
    number = getNumber(number,float)
    # Format the number to the specified number of decimal places
    number = f"{number:,.{decimalPoints}f}"   
    # Return the currency with formatted number
    return f"{currency} {number}"

def convertFunctionString(text):
    """ extract function names from a string
        and returns it in a dictionary varible"""
    #check if text is a string
    if not isinstance(text,str):
        return text
    # Regular expression to extract the function name and parameters
    pattern = r"(\w+)\((.*?)\)"
    match = re.match(pattern, text)    
    if match:
        # Extract the function name
        funcName = match.group(1)
        # Extract and split the parameters
        params = match.group(2).split(',')        
        # Create the dictionary
        result = {"name": funcName}
        for i, param in enumerate(params, start=1):
            param = extractValueFromString(param)
            result[f"param{i}"] = param
        return result
    else:
        return text

def extractValueFromString(value):
    """ try and extract value from a string
        This can return an int, float or a string"""
    value = value.strip()
    # Check if the parameter is a number (integer or float)
    if value.isdigit():
        # Convert to integer if it's an integer
        value = int(value)
    else:
        try:
            # Convert to float if it can be converted
            value = float(value)
        except ValueError:
            # Otherwise, keep it as a string
            value = str(value)
    print("Fn [extractValueFromString] => ",value, type(value))
    return value
