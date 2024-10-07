import msoffcrypto
import openpyxl
import os
import re
import getpass

TEMPLATE_FILE_NAME = "test/TEMPLATE.xlsx"
SOURCE_PATH = "test/"
# passwords perdata and saldata 

ERROR_SUCCESS = 0
ERROR_FILE_NOT_FOUND = 1

# create temp data file from source.
def createTempDataFile(sourceFile, password, tempFile):
    print("Fn: createTempDataFile")
    #check if the source file is there
    if( not os.path.exists(sourceFile)):
        print("Error: Source file not found")
        return ERROR_FILE_NOT_FOUND # error 
    with open(sourceFile, 'rb') as file:
        office_file = msoffcrypto.OfficeFile(file)
        office_file.load_key(password=password)  # Provide the password
        # Decrypt the file and save it as a temporary file
        with open(tempFile, 'wb') as decrypted_file:
            office_file.decrypt(decrypted_file)
    return ERROR_SUCCESS

# Remove temp files.
# the file list is a python dict, that has two elememnts.
# name: string delete: True/False
# return the error count
def removeTempFiles(fileList):
    print("Fn: removeTempFiles")
    errorCount = 0
    for files in fileList:
        if(files["delete"] & os.path.exists(files["name"])):
            os.remove(files["name"])
            print(f"Temporary files deleted.")
        else:
            print(f"ERROR: File not found.")
            print(files["name"])
            errorCount = errorCount + 1
    return errorCount

#Template column defs
TEMP_COL_INDEX = 0
TEMP_COL_NAME = 1
TEMP_COL_CONTENT = 2
TEMP_COL_LOC_X = 3
TEMP_COL_LOC_Y = 4
#Data content defs
TEMP_DATA_TYPE = 0
TEMP_DATA_FILE_NAME = 1
TEMP_DATA_IMD_TEXT = 1
TEMP_DATA_FILE_LOCKED = 2
TEMP_DATA_FILE_SHEET = 3
TEMP_DATA_FILE_COL_KEY = 4
TEMP_DATA_FILE_COL_DATA = 5

# Read template file, and load the data 
def loadTemplateData(templateFile,sheetName):
    # variable to return data
    textOverlayList = [] 
    #check if the file path is valid
    if( not os.path.exists(templateFile)):
        print("Error: Tempate file not found")
        return textOverlayList # error - empty
    #open template file, and load the sheet
    templateBook = openpyxl.load_workbook(templateFile)
    textOverlayDataSheet = templateBook[sheetName]
    print("Debug:",textOverlayDataSheet.dimensions, textOverlayDataSheet.max_row, textOverlayDataSheet.max_column)
    #go through every text overlay item
    for overlays in textOverlayDataSheet.rows:
        #Get the index to a string. 
        rowIndex = str(overlays[TEMP_COL_INDEX].value)
        # stop if we reach an empty cell
        if(None == rowIndex):
            break
        # the index has to be always numbers, skip others
        if( not rowIndex.isdigit()):
            print("Warning: skip Row Index : ", rowIndex)
            continue
        dataString = str(overlays[TEMP_COL_CONTENT].value)
        #user initiated end of loop.
        if("None" == dataString):
            print("Warning: User terminated at Index : ", rowIndex)
            break
        if(not (dataString.startswith('<') and dataString.endswith('>') and len(dataString) > 3)):
            print("Error: Data Error at Index : ",rowIndex)
            break
        #process the file data. Get all the data points to a list
        data = re.findall(r'<(.*?)>',dataString)
        #check the item 2, File locked
        if "!T" == data[TEMP_DATA_TYPE]:
            print("text")
            #Save notmal text data
            textOverlayList.append({"name": overlays[TEMP_COL_NAME].value, 
                                    "text":{ "string": data[TEMP_DATA_IMD_TEXT], 
                                            "x": overlays[TEMP_COL_LOC_X].value, 
                                            "y": overlays[TEMP_COL_LOC_Y].value}})
        elif "!F" == data[0]:
            print("file named : ",data[TEMP_DATA_FILE_NAME])
            isFileLocked = False
            if data[TEMP_DATA_FILE_LOCKED] == "LOCKED=1":
                isFileLocked = True
                #save the extended data
                textOverlayList.append({"name": overlays[TEMP_COL_NAME].value, 
                                        "text":{ "string": None, 
                                                "x": overlays[TEMP_COL_LOC_X].value, 
                                                "y": overlays[TEMP_COL_LOC_Y].value},
                                        "file" : {"name": data[TEMP_DATA_FILE_NAME], 
                                                  "isLocked": isFileLocked, 
                                                  "sheet": data[TEMP_DATA_FILE_SHEET], 
                                                  "primeryKey": data[TEMP_DATA_FILE_COL_KEY], 
                                                  "value": data[TEMP_DATA_FILE_COL_DATA]}})
        else:
            print("Error: undefined overlay type : ", data[0])
            break
    # Data store is done. return
    print("Fn: loadTemplateData")
    return textOverlayList

def getFileList(textOverlayList):
    fileNameList = []
    #go through each overlay
    for testOverlay in textOverlayList:
        if "file" in testOverlay:
            filedata = testOverlay["file"]
            filename = filedata["name"]
            print("There is a file ", filename)
            if filename not in fileNameList:
                fileNameList.append(filename)
    return fileNameList


def OpenAllDataFiles(overlayDataList):
    dataFileList =[]
    return dataFileList


# main program

#load the template
textOverlayList =  loadTemplateData(TEMPLATE_FILE_NAME,"Overlay")

#get the locked files to be used as source 
SourceFileList = getFileList(textOverlayList)

